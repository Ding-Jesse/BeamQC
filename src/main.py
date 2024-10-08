from __future__ import annotations
from multiprocessing.spawn import prepare
import time
import multiprocessing
# from AutocadApi import read_plan,read_beam,write_beam,write_plan,error
from src.plan_to_beam import run_plan, run_beam, output_error_list, write_beam, write_plan, write_result_log
from werkzeug.utils import secure_filename
import os
import src.plan_to_col as plan_to_col
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from src.save_temp_file import save_pkl, read_temp
from utils.connect_db import get_db, add_error_log


def main_functionV3(beam_filenames,
                    plan_filenames,
                    beam_new_filename,
                    plan_new_filename,
                    project_name,
                    output_directory,
                    layer_config,
                    sizing,
                    mline_scaling,
                    client_id,
                    plan_pkl="",
                    beam_pkl="",
                    plan_drawing_unit='cm',
                    beam_drawing_unit='cm') -> str:
    plan_result_dict = None
    beam_result_dict = None

    task_name = project_name
    date = time.strftime("%Y%m%d", time.localtime())
    data_excel_file = os.path.join(
        output_directory, f'{project_name}_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_梁Check_結果.xlsx')

    multiprocessing.freeze_support()
    pool = multiprocessing.Pool()

    res_plan = []
    res_beam = []
    set_plan = set()
    dic_plan = {}
    set_beam = set()
    dic_beam = {}

    for plan_filename in plan_filenames:
        res_plan.append(pool.apply_async(run_plan, (plan_filename,
                                                    plan_new_filename,
                                                    layer_config,
                                                    sizing,
                                                    mline_scaling,
                                                    date,
                                                    client_id,
                                                    plan_drawing_unit,
                                                    plan_pkl)))
    for beam_filename in beam_filenames:
        res_beam.append(pool.apply_async(run_beam, (beam_filename,
                                                    layer_config,
                                                    sizing,
                                                    client_id,
                                                    beam_drawing_unit,
                                                    beam_pkl)))

    plan_drawing = 0
    if len(plan_filenames) == 1:
        plan_drawing = 1
    beam_drawing = 0
    if len(beam_filenames) == 1:
        beam_drawing = 1

    mline_error_list = []
    plan_cad_data_list = []
    # try:
    for plan in res_plan:
        plan = plan.get()
        if plan:
            set_plan = set_plan | plan[0]
            if plan_drawing:
                dic_plan = plan[1]
            plan_mline_error_list = plan[2]
            plan_cad_data = plan[3]
            mline_error_list.extend(plan_mline_error_list)
            plan_cad_data_list.append(plan_cad_data)
        else:
            end = time.time()

    for beam in res_beam:
        beam = beam.get()
        if beam:
            set_beam = set_beam | beam[0]
            if beam_drawing:
                dic_beam = beam[1]
        else:
            end = time.time()

    plan_error_list = write_plan(plan_filename=plan_filename,
                                 plan_new_filename=plan_new_filename,
                                 set_plan=set_plan,
                                 set_beam=set_beam,
                                 dic_plan=dic_plan,
                                 date=date,
                                 drawing=plan_drawing,
                                 mline_scaling=mline_scaling,
                                 client_id=client_id)

    plan_error_list.extend(plan_mline_error_list)
    plan_error_counter, plan_result_dict = output_error_list(error_list=plan_error_list,
                                                             title_text='XS-BEAM',
                                                             set_item=set_plan,
                                                             cad_data=plan_cad_data_list)

    beam_error_list = write_beam(beam_filename=beam_filename,
                                 beam_new_filename=beam_new_filename,
                                 set_plan=set_plan,
                                 set_beam=set_beam,
                                 dic_beam=dic_beam,
                                 date=date,
                                 drawing=beam_drawing,
                                 client_id=client_id)
    beam_error_counter, beam_result_dict = output_error_list(error_list=beam_error_list,
                                                             title_text='XS-PLAN',
                                                             set_item=set_beam)
    end = time.time()

    # save_pkl({
    #     'plan_result_dict': plan_result_dict,
    #     'plan_error_counter': plan_error_counter,
    #     'beam_result_dict': beam_result_dict,
    #     'beam_error_counter': beam_error_counter,
    # }, tmp_file=os.path.join(output_directory, f'{project_name}-result.pkl'))

    output_data = write_result_log(task_name=task_name,
                                   plan_result=plan_result_dict,
                                   beam_result=beam_result_dict
                                   )
    for sheet_name, df_list in output_data.items():
        OutputExcel(df_list=df_list,
                    df_spacing=1,
                    file_path=data_excel_file,
                    sheet_name=sheet_name)
    Upload_Error_log(data={
        'date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        'beam_filenames': beam_filenames,
        'beam_new_filename': beam_new_filename,
        'plan_filenames': plan_filenames,
        'plan_new_filename': plan_new_filename,
        'project_name': task_name,
        'output_directory': output_directory,
        "layer_config": layer_config,
        'sizing': sizing,
        'mline_scaling': mline_scaling,
        'client_id': client_id,
        'plan error rate': plan_result_dict['summary'] if plan_result_dict is not None else None,
        'beam error rate': beam_result_dict['summary'] if beam_result_dict is not None else None
    }, collection_name='Beam Check Log')
    # except Exception as ex:
    #     status = 'error'
    #     Upload_Error_log(data={
    #         'date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
    #         'beam_filenames': beam_filenames,
    #         'beam_new_filename': beam_new_filename,
    #         'plan_filenames': plan_filenames,
    #         'plan_new_filename': plan_new_filename,
    #         'project_name': task_name,
    #         'output_directory': output_directory,
    #         "layer_config": layer_config,
    #         'sizing': sizing,
    #         'mline_scaling': mline_scaling,
    #         'client_id': client_id,
    #         'status': status,
    #         'plan error rate': plan_result_dict['summary'] if plan_result_dict is not None else None,
    #         'beam error rate': beam_result_dict['summary'] if beam_result_dict is not None else None,
    #         'error':ex,
    #     }, collection_name='Beam Check Log')
    return os.path.basename(data_excel_file)


def main_col_function(col_filenames,
                      plan_filenames,
                      col_new_filename,
                      plan_new_filename,
                      layer_config,
                      output_directory,
                      project_name,
                      client_id,
                      plan_drawing_unit='cm',
                      column_drawing_unit='cm',
                      column_bottom_line=1,
                      exclude_string: list = None,
                      plan_pkl: str = "",
                      col_pkl: str = ""):
    '''
    Args:
        layer_config:{text_layer,line_layer,block_layer,floor_layer,col_layer}
    '''
    if exclude_string is None:
        exclude_string = []
    start = time.time()
    plan_result_dict = None
    col_result_dict = None
    status = 'progress'
    # text_layer,line_layer,block_layer,floor_layer,col_layer,
    data_excel_file = os.path.join(
        output_directory, f'{project_name}_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_柱Check_結果.xlsx')

    date = time.strftime("%Y-%m-%d", time.localtime())

    multiprocessing.freeze_support()
    pool = multiprocessing.Pool()
    res_plan = []
    res_col = []
    set_plan = set()
    dic_plan = {}
    set_col = set()
    dic_col = {}
    plan_dwg_file = None
    col_dwg_file = None
    for plan_dwg_file in plan_filenames:
        res_plan.append(pool.apply_async(plan_to_col.run_plan, (plan_dwg_file,
                                                                layer_config,
                                                                client_id,
                                                                plan_drawing_unit,
                                                                plan_pkl)))

    for col_dwg_file in col_filenames:
        res_col.append(pool.apply_async(plan_to_col.run_col, (col_dwg_file,
                                                              layer_config,
                                                              client_id,
                                                              column_drawing_unit,
                                                              column_bottom_line,
                                                              exclude_string,
                                                              col_pkl)))

    plan_drawing = 0
    if len(plan_filenames) == 1:
        plan_drawing = 1
    col_drawing = 0
    if len(col_filenames) == 1:
        col_drawing = 1

    plan_block_error_list = []
    plan_block_match_result_list = []

    for plan in res_plan:
        plan = plan.get()
        if plan:
            set_plan = set_plan | plan[0]
            if plan_drawing:
                dic_plan = plan[1]
            block_error_list = plan[2]
            block_match_result_list = plan[3]
            plan_block_error_list.extend(block_error_list)
            plan_block_match_result_list.extend(block_match_result_list)
        else:
            end = time.time()
            # plan_to_col.write_result_log(
            #     excel_file, task_name, '', '', f'{round(end - start, 2)}(s)', time.strftime("%Y-%m-%d %H:%M", time.localtime()), 'failed')
            # return

    for col in res_col:
        col = col.get()
        if col:
            set_col = set_col | col[0]
            if col_drawing:
                dic_col = col[1]
        else:
            end = time.time()
            # plan_to_col.write_result_log(
            #     excel_file, task_name, '', '', f'{round(end - start, 2)}(s)', time.strftime("%Y-%m-%d %H:%M", time.localtime()), 'failed')
            # return
    # try:
    plan_result = plan_to_col.write_plan(plan_dwg_file,
                                         plan_new_filename,
                                         set_plan,
                                         set_col,
                                         dic_plan,
                                         date,
                                         plan_drawing,
                                         block_match=plan_block_match_result_list,
                                         client_id=client_id)
    col_result = plan_to_col.write_col(col_dwg_file,
                                       col_new_filename,
                                       set_plan,
                                       set_col,
                                       dic_col,
                                       date,
                                       col_drawing,
                                       client_id=client_id)

    plan_result_dict, col_result_dict, excel_data = plan_to_col.write_result_log(plan_error_list=plan_result,
                                                                                 col_error_list=col_result,
                                                                                 set_plan=set_plan,
                                                                                 set_col=set_col,
                                                                                 block_error_list=plan_block_error_list,
                                                                                 block_match_list=block_match_result_list
                                                                                 )
    for sheet_name, df_list in excel_data.items():
        OutputExcel(df_list=df_list,
                    df_spacing=1,
                    file_path=data_excel_file,
                    sheet_name=sheet_name)
    Upload_Error_log(data={
        'date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        'col_filenames': col_filenames,
        'col_new_filename': col_new_filename,
        'plan_filenames': plan_filenames,
        'plan_new_filename': plan_new_filename,
        'project_name': project_name,
        'output_directory': output_directory,
        "layer_config": layer_config,
        'client_id': client_id,
        'status': status,
        'plan error rate': plan_result_dict['summary'] if plan_result_dict is not None else None,
        'col error rate': col_result_dict['summary'] if col_result_dict is not None else None
    }, collection_name='Column Check Log')
    # except Exception as ex:
    #     print(ex)
    #     status = 'error'
    # finally:
    #     Upload_Error_log(data={
    #         'date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
    #         'col_filenames': col_filenames,
    #         'col_new_filename': col_new_filename,
    #         'plan_filenames': plan_filenames,
    #         'plan_new_filename': plan_new_filename,
    #         'project_name': project_name,
    #         'output_directory': output_directory,
    #         "layer_config": layer_config,
    #         'client_id': client_id,
    #         'status': status,
    #         'plan error rate': plan_result_dict['summary'] if plan_result_dict is not None else None,
    #         'beam error rate': col_result_dict['summary'] if col_result_dict is not None else None
    #     }, collection_name='Column Check Log')
    #     if status == 'error':
    #         raise Exception
    return os.path.basename(data_excel_file)


def storefile(file, file_directory, file_new_directory, project_name):
    filename_beam = secure_filename(file.filename)
    check_beam_name = os.path.splitext(filename_beam)
    if check_beam_name[1] == '':
        filename_beam = f'temp.{check_beam_name[0]}'
    # print(f'file:{file.filename} file:{filename_beam}')
    save_file = os.path.join(file_directory, f'{project_name}-{filename_beam}')
    file_new_name = os.path.join(
        file_new_directory, f'{project_name}_MARKON-{filename_beam}')
    file.save(save_file)
    file_ok = True
    return file_ok, file_new_name, save_file


def OutputExcel(df_list: list[pd.DataFrame], file_path, sheet_name, auto_fit_columns=[], auto_fit_rows=[], columns_list=[], rows_list=[], df_spacing=0):
    if os.path.exists(file_path):
        # book = load_workbook(file_path)
        writer = pd.ExcelWriter(
            file_path, engine='openpyxl', mode="a", if_sheet_exists="overlay")
        # writer.book = book
        # sheet = book[sheet_name]
        # sheet.column_dimensions['A'] =ColumnDimension(sheet,'L',bestFit=True)
    else:
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    row = 0
    for df in df_list:
        df.to_excel(writer, sheet_name=sheet_name, startrow=row)
        row += len(df.index) + df_spacing
    writer.close()

    # book = load_workbook(file_path)
    # writer = pd.ExcelWriter(file_path, engine='openpyxl',mode="a",if_sheet_exists="overlay")
    with pd.ExcelWriter(file_path, engine='openpyxl', mode="a", if_sheet_exists="overlay") as writer:
        # writer.book = book
        if os.path.exists(file_path) and len(auto_fit_columns) > 0:
            AutoFit_Columns(writer.book[sheet_name],
                            auto_fit_columns, auto_fit_rows)
        if os.path.exists(file_path) and len(columns_list) > 0:
            Decorate_Worksheet(
                writer.book[sheet_name], columns_list, rows_list)
    # writer.close()
    return file_path


def Decorate_Worksheet(sheet: Worksheet, columns_list: list, rows_list: list):
    for i in columns_list:
        for j in rows_list:
            sheet.cell(j, i).alignment = Alignment(
                vertical='center', wrap_text=True, horizontal='center')
            sheet.cell(j, i).font = Font(name='Calibri')
            if sheet.cell(j, i).value == 'NG.':
                sheet.cell(j, i).fill = PatternFill(
                    "solid", start_color='00FF0000')


def AutoFit_Columns(sheet: Worksheet, auto_fit_columns: list, auto_fit_rows: list):
    for i in auto_fit_columns:
        sheet.column_dimensions[get_column_letter(i)].width = 80
    for i in auto_fit_rows:
        sheet.row_dimensions[i].height = 20
    for i in auto_fit_rows:
        for j in auto_fit_columns:
            sheet.cell(i, j).alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')


def Add_Row_Title(file_path: str, sheet_name: str, i: int, j: int, title_text: str, font_size=12):
    # book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl',
                            mode="a", if_sheet_exists="overlay")
    # writer.book = book
    sheet = writer.book[sheet_name]
    sheet.cell(i, j).value = title_text
    sheet.cell(i, j).alignment = Alignment(
        vertical='center', wrap_text=True, horizontal='center')
    sheet.cell(i, j).font = Font(name='Calibri', size=font_size)
    writer.close()


def Output_Config(project_name: str, layer_config: dict, file_new_directory: str):
    with open(os.path.join(file_new_directory, f'{project_name}_layer_config.txt'), 'w') as f:
        f.write(str(layer_config))


def Upload_Error_log(data, collection_name="Log", uri=None):
    # uri = os.environ['MONGO_URL'].replace('"', '')
    db = get_db('RcCheck', uri=uri)
    add_error_log(db,
                  data=data,
                  collection_name=collection_name)


def GetAllFiles(mypath: str):
    import glob
    from os.path import isfile, join
    return glob.glob(join(mypath, "*.dwg"))


if __name__ == '__main__':
    output_directory = r"D:\Desktop\BeamQC\TEST\2024-0605"
    task_name = project_name = "0524-temp"
    # Upload_Error_log([{1: 1}, {
    #                  2: 2}], uri="mongodb+srv://ghjk85692012:MGI3hjs341a7kNWq@cluster0.glspevs.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0")
    # data = read_temp(tmp_file=r'TEST\2024-0522\temp-result.pkl')
    # plan_result_dict = data['plan_result_dict']
    # beam_result_dict = data['beam_result_dict']
    # data_excel_file = os.path.join(
    #     output_directory, f'{project_name}-結果-test2.xlsx')
    # output_data = write_result_log(excel_file='',
    #                                task_name=task_name,
    #                                plan_result=plan_result_dict,
    #                                beam_result=beam_result_dict,
    #                                runtime=f'0(s)',
    #                                date=time.strftime(
    #                                    "%Y-%m-%d %H:%M", time.localtime()),
    #                                other='none')
    # for sheet_name, df_list in output_data.items():
    #     OutputExcel(df_list=df_list,
    #                 df_spacing=1,
    #                 file_path=data_excel_file,
    #                 sheet_name=sheet_name)

    beam_filenames = [
        r'D:\Desktop\BeamQC\TEST\2024-0524\2024-05-23-09-53_temp-XS-BEAM.dwg']
    # sys.argv[2] # XS-PLAN的路徑
    plan_filenames = [
        r'D:\Desktop\BeamQC\TEST\2024-0524\2024-05-23-09-53_temp-XS-PLAN.dwg']
    # sys.argv[3] # XS-BEAM_new的路徑
    beam_new_filename = f"{output_directory}/0524-XS-BEAM_new2.dwg"
    # sys.argv[4] # XS-PLAN_new的路徑
    plan_new_filename = f"{output_directory}/0524-XS-PLAN_new2.dwg"

    column_filenames = []

    # # 在beam裡面自訂圖層
    text_layer = ['S-RC']  # sys.argv[7]

    # 在plan裡面自訂圖層
    block_layer = ['DwFm', '0', 'DETPOINTS']  # sys.argv[8] # 框框的圖層
    floor_layer = ['S-TITLE']  # sys.argv[9] # 樓層字串的圖層
    size_layer = ['S-TEXT']  # sys.argv[12] # 梁尺寸字串圖層
    big_beam_layer = ['S-RCBMG']  # 大樑複線圖層
    big_beam_text_layer = ['S-TEXTG']  # 大樑文字圖層
    sml_beam_layer = ['S-RCBMB']  # 小梁複線圖層
    sml_beam_text_layer = ['S-TEXTB']  # 小梁文字圖層
    task_name = '0524-temp'  # sys.argv[13]

    progress_file = './result/tmp'  # sys.argv[14]

    sizing = 1  # 要不要對尺寸
    mline_scaling = 1  # 要不要對複線寬度

    # plan_file = './result/plan.txt' # plan.txt的路徑
    # beam_file = './result/beam.txt' # beam.txt的路徑
    # excel_file = './result/result_log.xlsx' # result_log.xlsx的路徑

    # date = time.strftime("%Y-%m-%d", time.localtime())
    layer_config = {
        # 'line_layer':line_layer,
        'text_layer': text_layer,
        'block_layer': block_layer,
        'floor_layer': floor_layer,
        'big_beam_layer': big_beam_layer,
        'big_beam_text_layer': big_beam_text_layer,
        'sml_beam_layer': sml_beam_layer,
        'size_layer': size_layer,
        'sml_beam_text_layer': sml_beam_text_layer
        # 'col_layer':col_layer
    }

    # col_filenames = [
    #     r'D:\Desktop\BeamQC\TEST\2023-1017\2023-10-17-09-25temp-XS-SCOL.dwg']
    # # print(col_filename.split(','))
    # # col_filename = r'D:\Desktop\BeamQC\TEST\2023-0303\2023-0301 左棟主配筋圖.dwg'
    # # sys.argv[2] # XS-PLAN的路徑
    # plan_filenames = [
    #     r'D:\Desktop\BeamQC\TEST\2023-1017\2023-10-17-09-25temp-XS-PLAN_.dwg']
    # # sys.argv[3] # XS-COL_new的路徑
    # col_new_filename = r'D:\Desktop\BeamQC\TEST\2023-1017\XS_new.dwg'
    # # sys.argv[4] # XS-PLAN_new的路徑
    # plan_new_filename = r'D:\Desktop\BeamQC\TEST\2023-1017\XS-PLAN_col_new.dwg'
    # # sys.argv[5] # 柱配筋結果
    # result_file = r'D:\Desktop\BeamQC\TEST\2023-1017\column.txt'

    # # 在col裡面自訂圖層
    # layer_config = {
    #     'text_layer': ['S-TEXT'],
    #     'line_layer': ['S-TABLE'],
    #     'block_layer': ['0', 'DwFm'],
    #     'floor_layer': ['S-TITLE'],
    #     'col_layer': ['S-TEXTC'],
    #     'size_layer': ['S-TEXT'],
    #     'table_line_layer': ['S-TABLE']
    # }
    # text_layer = 'SS'#sys.argv[6] # 文字的圖層
    # line_layer = 'S-TABLE'#sys.argv[7] # 線的圖層

    # 在plan裡面自訂圖層
    # block_layer = 'ss-dwfm'#sys.argv[8] # 圖框的圖層
    # floor_layer = 'S-TITLE'#sys.argv[9] # 樓層字串的圖層
    # col_layer = 'ss-dim'#sys.argv[10] # col的圖層

    # task_name = 'tmp'  # sys.argv[11]

    # progress_file = './result/tmp'  # sys.argv[12]
    # plan_file = './result/col_plan.txt'  # plan.txt的路徑
    # col_file = './result/col.txt'  # col.txt的路徑
    # excel_file = './result/result_log_col.xlsx'  # result_log.xlsx的路徑
    # main_col_function(col_filenames=col_filenames,
    #                   col_new_filename=col_new_filename,
    #                   plan_filenames=plan_filenames,
    #                   plan_new_filename=plan_new_filename,
    #                   result_file=result_file,
    #                   layer_config=layer_config,
    #                   task_name=task_name,
    #                   progress_file=progress_file,
    #                   client_id="0522-temp",
    #                   plan_pkl=r"TEST\2024-0522\2024-05-20-19-07P2023-05A 桃園龜山樂善安居14FB3-XS-PLAN_plan_to_col.pkl",
    #                   col_pkl=r"D:\Desktop\BeamQC\TEST\2024-0522\2024-05-21-14-57P2023-05A 桃園龜山樂善安居14FB3-XS-COL_col_set.pkl")

    # from collections import Counter
    main_functionV3(beam_filenames=beam_filenames,
                    beam_new_filename=beam_new_filename,
                    plan_filenames=plan_filenames,
                    plan_new_filename=plan_new_filename,
                    project_name=task_name,
                    output_directory=output_directory,
                    layer_config=layer_config,
                    sizing=sizing,
                    mline_scaling=mline_scaling,
                    client_id="2024-0605",
                    plan_pkl=r'TEST\2024-0605\2024-06-14-14-57_2024-0614 佳元2-XS-PLAN_plan_set.pkl',
                    beam_pkl=r'TEST\2024-0605\2024-06-14-14-57_2024-0614 佳元2-XS-BEAM_beam_set.pkl')

    # Column Test
    # layer_config = {
    #     'text_layer': ['S-TEXT'],
    #     'line_layer': ['S-TABLE'],
    #     'block_layer': ['0', 'DwFm', 'DEFPOINTS'],
    #     'floor_layer': ['S-TITLE'],
    #     'col_layer': ['S-TEXTC'],
    #     'size_layer': ['S-TEXT'],
    #     'table_line_layer': ['S-TABLE'],
    #     'column_block_layer': ['S-COL']
    # }
    # col_filenames = [
    #     r'D:\Desktop\BeamQC\TEST\2024-0528\2024-05-28-11-50_temp-A.dwg',
    #     r'D:\Desktop\BeamQC\TEST\2024-0528\2024-05-28-11-50_temp-B.dwg']
    # plan_filenames = [
    #     r'D:\Desktop\BeamQC\TEST\2024-0528\2024-05-28-11-50_temp-XS-PLAN.dwg']
    # col_new_filename = r'D:\Desktop\BeamQC\TEST\2024-0528\2024-05-24-10-02_temp-XS-COL2.dwg'
    # plan_new_filename = r'D:\Desktop\BeamQC\TEST\2024-0528\2024-05-24-10-02_temp-XS-PLAN2.dwg'
    # main_col_function(
    #     col_filenames=col_filenames,
    #     plan_filenames=plan_filenames,
    #     col_new_filename=col_new_filename,
    #     plan_new_filename=plan_new_filename,
    #     output_directory=output_directory,
    #     project_name=project_name,
    #     layer_config=layer_config,
    #     client_id="0529-col",
    #     plan_pkl=r'TEST\2024-0605\2024-06-14-14-57_2024-0614 佳元2-XS-PLAN_plan_to_col.pkl',
    #     col_pkl=r'TEST\2024-0605\2024-06-14-14-57_2024-0614 佳元2-XS-COL_col_set.pkl'
    # )
