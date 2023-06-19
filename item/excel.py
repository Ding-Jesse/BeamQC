import pandas as pd
# import excel2img
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side


def AddExcelDataBar(workbook_path: str, sheet_name: str, start_col: int, start_row: int, end_col: int, end_row: int):
    # book = load_workbook(workbook_path)
    writer = pd.ExcelWriter(workbook_path, engine='openpyxl',
                            mode="a", if_sheet_exists="overlay")
    # writer.book = book
    sheet = writer.book[sheet_name]

    condition_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
    rows = sheet[condition_range]
    for row in rows:
        for cell in row:
            cell.number_format = '0%'
    rule = ColorScaleRule(start_type='min', start_color='63BE7B',
                          mid_type='percentile', mid_value=50, mid_color='FFEB84',
                          end_type='max', end_color='F8696B')
    sheet.conditional_formatting.add(condition_range, rule)
    writer.close()
    # ExportExportRangeToImage(workbook_path,r'D:\Desktop\BeamQC\assets\ratio_image.png',sheet_name,None)


def AddBorderLine(workbook_path: str, sheet_name: str, start_col: int, start_row: int, end_col: int, end_row: int, step_row: int, step_col: int):
    # book = load_workbook(workbook_path)
    writer = pd.ExcelWriter(workbook_path, engine='openpyxl',
                            mode="a", if_sheet_exists="overlay")
    # writer.book = book
    sheet = writer.book[sheet_name]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    double = Side(border_style="double", color="000000")
    thin = Side(border_style="thin", color="000000")
    cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
    rows = sheet[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border
    for row in range(start_row, end_row, step_row):
        for rows in sheet.iter_rows(min_row=row, max_row=row, min_col=None):
            for cell in rows:
                cell.border = Border(top=double, left=thin,
                                     right=thin, bottom=thin)
    # for col in range(start_col,end_col,step_col):
    #     for rows in sheet.iter_rows(min_col=col):
    #         for cell in rows:
    #             cell.border = Border(right=double)
    # for col in range(start_col,end_col,step_col):
    #     for row in range(start_row,end_row,step_row):
    #         condition_range = f'{get_column_letter(col)}{row}:{get_column_letter(col + step_col)}{row + step_row}'
    #         range_cell = sheet[condition_range]
    #         range_cell.border = Border(outline=double)
    writer.close()


def ExportExportRangeToImage(xlsx_filename: str, output_image: str, sheet_name: str, cell_range: str):
    '''
    ### Save as PNG the range of used cells in test.xlsx on page named "Sheet1"
    - excel2img.export_img("test.xlsx", "test.png", "Sheet1", None)
    ### Save as BMP the range B2:C15 in test.xlsx on page named "Sheet2"
    - excel2img.export_img("test.xlsx", "test.bmp", "", "Sheet2!B2:C15")
    '''
    # excel2img.export_img(fn_excel= xlsx_filename,
    #                      fn_image= output_image,
    #                      page=sheet_name,
    #                      _range = None)
