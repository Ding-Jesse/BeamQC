from __future__ import annotations
import time
import pythoncom
import win32com.client
import re
import os
import save_temp_file
from math import sqrt,pow
from rebar import RebarInfo
from plan_to_beam import turn_floor_to_float, turn_floor_to_string, turn_floor_to_list, floor_exist, vtFloat, error
from column import Column,Floor
from beam_count import vtPnt
from column_scan import column_check,create_column_scan
import pandas as pd
import numpy as np
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from multiprocessing.pool import ThreadPool as Pool
import multiprocessing as mp

slash_pattern = r'(.+)[~|-](.+)' #~
commom_pattern = r'(,)|(、)'
multi = True
def read_column_cad(column_filename,layer_config:dict[list]):
    layer_list = [layer for key,layer in layer_config.items()]
    # line_layer = layer_config['line_layer']
    error_count = 0
    pythoncom.CoInitialize()
    # Step 1. 打開應用程式
    flag = 0
    while not flag and error_count <= 10:
        try:
            wincad_column = win32com.client.Dispatch("AutoCAD.Application")
            flag = 1
        except Exception as e:
            error_count += 1
            time.sleep(5)
            error(f'read_beam error in step 1: {e}, error_count = {error_count}.')

    # Step 2. 匯入檔案
    flag = 0
    while not flag and error_count <= 10:
        try:
            doc_column = wincad_column.Documents.Open(column_filename)
            flag = 1
        except Exception as e:
            error_count += 1
            time.sleep(5)
            error(f'read_beam error in step 2: {e}, error_count = {error_count}.')

    # Step 3. 匯入modelspace
    flag = 0
    while not flag and error_count <= 10:
        try:
            msp_column = doc_column.Modelspace
            flag = 1
        except Exception as e:
            error_count += 1
            time.sleep(5)
            error(f'read_beam error in step 3: {e}, error_count = {error_count}.')

    # 在這之後就沒有while迴圈了，所以錯超過10次就出去
    if error_count > 10:
        try:
            doc_column.Close(SaveChanges=False)
        except:
            pass
        return False
    # flag = 0
    # # Step 4. 解鎖所有圖層 -> 不然不能刪東西
    # while not flag and error_count <= 10:
    #     try:
    #         layer_count = doc_column.Layers.count

    #         for x in range(layer_count):
    #             layer = doc_column.Layers.Item(x)
    #             layer.Lock = False
    #         flag = 1
    #     except Exception as e:
    #         error_count += 1
    #         time.sleep(5)
    #         msp_column  = doc_column.Modelspace
    #         error(f'read_col error in step 4: {e}, error_count = {error_count}.')
    # progress('柱配筋圖讀取進度 4/10', progress_file)

    # Step 5. 遍歷所有物件 -> 炸圖塊  
    # flag = 0
    # while not flag and error_count <= 10:
    #     try:
    #         count = 0
    #         total = msp_col.Count
    #         # layer_list = [text_layer, line_layer]
    #         for object in msp_col:
    #             count += 1
    #             if object.EntityName == "AcDbBlockReference" and object.Layer in layer_list:
    #                 object.Explode()
    #             if object.Layer not in layer_list:
    #                 object.Delete()
    #         flag = 1
    #     except Exception as e:
    #         error_count += 1
    #         time.sleep(5)
    #         try:
    #             msp_col = doc_column.Modelspace
    #         except:
    #             pass
    #         error(f'read_col error in step 5: {e}, error_count = {error_count}.')
    # progress('柱配筋圖讀取進度 5/10', progress_file)

    # Step 6. 重新匯入modelspace
    # flag = 0
    # while not flag and error_count <= 10:
    #     try:
    #         msp_column = doc_column.Modelspace
    #         flag = 1
    #     except Exception as e:
    #         error_count += 1
    #         time.sleep(5)
    #         error(f'read_col error in step 6: {e}, error_count = {error_count}.')
    return msp_column,doc_column

def sort_col_cad(msp_column,layer_config,temp_file):
    text_layer = list(layer_config['text_layer'])
    line_layer = list(layer_config['line_layer'])
    rebar_text_layer = list(layer_config['rebar_text_layer'])
    rebar_layer = list(layer_config['rebar_layer'])
    tie_layer = list(layer_config['tie_layer'])
    tie_text_layer = list(layer_config['tie_text_layer'])
    column_rc_layer = list(layer_config['column_rc_layer'])
    coor_to_floor_set = set() # set(coor, floor)
    coor_to_col_set = set() # set(coor, col)
    coor_to_size_set = set() # set(coor, size)
    coor_to_floor_line_list = [] # (橫線y座標, start, end)
    coor_to_col_line_list = [] # (縱線x座標, start, end)
    coor_to_rebar_text_list = []
    coor_to_rebar_list = []
    coor_to_tie_text_list = []
    coor_to_tie_list = []
    coor_to_section_list = []
    flag = 0
    error_count = 0
    while not flag and error_count <= 10:
        # try:
        count = 0
        total = msp_column.Count
        # progress(f'柱配筋圖上共有{total}個物件，大約運行{int(total / 9000) + 1}分鐘，請耐心等候', progress_file)
        for object in msp_column:
            count += 1
            # if count % 1000 == 0:
            #     progress(f'柱配筋圖已讀取{count}/{total}個物件', progress_file)
            print(f'{object.Layer}:{object.ObjectName}')
            if object.Layer in tie_layer:
                # print(f'{object.Layer}:{object.ObjectName}')
                if object.ObjectName == "AcDbPolyline":
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    coor_to_tie_list.append((coor1,coor2))
            if object.Layer in tie_text_layer:
                if object.ObjectName == "AcDbText":
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    coor_to_tie_text_list.append(((coor1,coor2),object.TextString))
            if object.Layer in rebar_text_layer and object.ObjectName == "AcDbText":
                if re.match(r'\d+.#\d+',object.TextString):
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    coor_to_rebar_text_list.append(((coor1,coor2),object.TextString))
            if object.Layer in rebar_layer:
                if object.ObjectName == "AcDbCircle":
                    coor1 = (round(object.Center[0],2),round(object.Center[1],2))
                    coor_to_rebar_list.append((coor1,'circle'))
                # if object.ObjectName == "AcDbPolyline":
                #     coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                #     coor_to_rebar_list.append(coor1)
                if object.ObjectName == "AcDbBlockReference":
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor_to_rebar_list.append((coor1,object.Name))

            if object.Layer in text_layer and object.ObjectName == "AcDbText": 
                if object.TextString[0] == 'C' and len(object.TextString) <= 7:
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    match_obj = re.search(slash_pattern,object.TextString)
                    if match_obj:
                        suffix_index = re.search(r'(\D+)\d+(\D+)',match_obj.group(1))
                        first_column = re.search(r'\d+',match_obj.group(1))
                        last_column = re.search(r'\d+',match_obj.group(2))
                        for column_number in range(first_column,last_column):
                            temp_string = f'{suffix_index.group(1)}{column_number}{suffix_index.group(2)}'
                            coor_to_col_set.add(((coor1, coor2), temp_string))
                    elif re.search(commom_pattern,object.TextString):
                        sep = re.search(commom_pattern,object.TextString).group(1)
                        for column_text in object.TextString.split(sep):
                            coor_to_col_set.add(((coor1, coor2), column_text))
                    else:
                        coor_to_col_set.add(((coor1, coor2), object.TextString))

                elif 'x' in object.TextString or 'X' in object.TextString:
                    size = object.TextString.replace('X', 'x')
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    coor_to_size_set.add(((coor1, coor2), size))
                elif ('F' in object.TextString or 'B' in object.TextString or 'R' in object.TextString) and 'O' not in object.TextString: # 可能有樓層
                    floor = object.TextString
                    if '_' in floor: # 可能有B_6F表示B棟的6F
                        floor = floor.split('_')[1]
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    if '~' in floor:
                        match_obj = re.search(r'(.+)[~](.+)',floor)
                        first_floor = int(turn_floor_to_float(match_obj.group(1)))
                        last_floor = int(turn_floor_to_float(match_obj.group(2)))
                        for floor_float in range(first_floor,last_floor + 1):
                            coor_to_floor_set.add(((coor1, coor2), turn_floor_to_string(floor_float)))
                    elif re.search(commom_pattern,floor):
                        sep = re.search(commom_pattern,object.TextString).group(1)
                        for floor_float in floor.split(sep):
                            coor_to_floor_set.add(((coor1, coor2), turn_floor_to_string(floor_float)))
                    else:   
                        if turn_floor_to_float(floor):
                            coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                            coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                            floor = turn_floor_to_float(floor)
                            floor = turn_floor_to_string(floor)
                            coor_to_floor_set.add(((coor1, coor2), floor))
            
            elif object.Layer in line_layer:
                coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                if coor1[0] == coor2[0]:
                    coor_to_col_line_list.append((coor1[0], min(coor1[1], coor2[1]), max(coor1[1], coor2[1])))
                elif coor1[1] == coor2[1]:
                    coor_to_floor_line_list.append((coor1[1], min(coor1[0], coor2[0]), max(coor1[0], coor2[0])))
            if object.Layer in column_rc_layer:
                if object.ObjectName == "AcDbPolyline":
                    coor1 = (round(object.GetBoundingBox()[0][0], 2), round(object.GetBoundingBox()[0][1], 2))
                    coor2 = (round(object.GetBoundingBox()[1][0], 2), round(object.GetBoundingBox()[1][1], 2))
                    coor_to_section_list.append((coor1,coor2))
        flag = 1
        coor_to_col_line_list.sort(key = lambda x: x[0])
        coor_to_floor_line_list.sort(key = lambda x: x[0])
        # except Exception as e:
        #     error_count += 1
        #     time.sleep(5)
        #     error(f'read_col error in step 7: {e}, error_count = {error_count}.')
    if multi:
        return{'coor_to_col_set':coor_to_col_set,
                'coor_to_size_set':coor_to_size_set,
                'coor_to_floor_set': coor_to_floor_set,
                'coor_to_col_line_list':coor_to_col_line_list,
                'coor_to_floor_line_list':coor_to_floor_line_list,
                'coor_to_rebar_text_list':coor_to_rebar_text_list,
                'coor_to_rebar_list':coor_to_rebar_list,
                'coor_to_tie_text_list':coor_to_tie_text_list,
                'coor_to_tie_list':coor_to_tie_list,
                'coor_to_section_list':coor_to_section_list
            }
    save_temp_file.save_pkl({'coor_to_col_set':coor_to_col_set,
                    'coor_to_size_set':coor_to_size_set,
                    'coor_to_floor_set': coor_to_floor_set,
                    'coor_to_col_line_list':coor_to_col_line_list,
                    'coor_to_floor_line_list':coor_to_floor_line_list,
                    'coor_to_rebar_text_list':coor_to_rebar_text_list,
                    'coor_to_rebar_list':coor_to_rebar_list,
                    'coor_to_tie_text_list':coor_to_tie_text_list,
                    'coor_to_tie_list':coor_to_tie_list,
                    'coor_to_section_list':coor_to_section_list
                    },temp_file)

def cal_column_rebar(data={},output_folder = '',project_name = '',msp_column = None ,doc_column = None,floor_parameter_xlsx=''):
    output_txt =os.path.join(output_folder,f'{project_name}_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_rebar.txt')
    output_txt_2 =os.path.join(output_folder,f'{project_name}_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_rebar_floor.txt')
    excel_filename = (
        f'{output_folder}/'
        f'{project_name}_'
        f'{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_'
        f'Count.xlsx'
    )
    if not data:
        return
    coor_to_col_set = data['coor_to_col_set']
    coor_to_size_set = data['coor_to_size_set']
    coor_to_floor_set = data['coor_to_floor_set']
    coor_to_col_line_list = data['coor_to_col_line_list']
    coor_to_floor_line_list= data['coor_to_floor_line_list']
    coor_to_rebar_text_list= data['coor_to_rebar_text_list']
    coor_to_rebar_list= data['coor_to_rebar_list']
    coor_to_tie_text_list= data['coor_to_tie_text_list']
    coor_to_tie_list= data['coor_to_tie_list']
    coor_to_section_list = data['coor_to_section_list']
    new_coor_to_col_line_list = concat_col_to_grid(coor_to_col_set=coor_to_col_set,coor_to_col_line_list=coor_to_col_line_list)
    new_coor_to_floor_line_list = concat_floor_to_grid(coor_to_floor_set=coor_to_floor_set,coor_to_floor_line_list=coor_to_floor_line_list)
    # draw_grid_line(new_coor_to_floor_line_list=new_coor_to_floor_line_list,new_coor_to_col_line_list=new_coor_to_col_line_list,msp_beam=msp_column,doc_beam=doc_column)
    # output_column_list = concat_name_to_col_floor(coor_to_size_set=coor_to_size_set,new_coor_to_col_line_list=new_coor_to_col_line_list,new_coor_to_floor_line_list=new_coor_to_floor_line_list)
    output_column_list = get_size_from_section(new_coor_to_col_line_list=new_coor_to_col_line_list,new_coor_to_floor_line_list=new_coor_to_floor_line_list
                                               ,coor_to_section_list=coor_to_section_list,coor_to_size_set=coor_to_size_set)
    combine_col_rebar(column_list=output_column_list,coor_to_rebar_list=coor_to_rebar_list,coor_to_rebar_text_list=coor_to_rebar_text_list)
    combine_col_tie(column_list=output_column_list,coor_to_tie_list=coor_to_tie_list,coor_to_tie_text_list=coor_to_tie_text_list)
    floor_list = floor_parameter(column_list=output_column_list,floor_parameter_xlsx=floor_parameter_xlsx)
    sort_floor_column(floor_list=floor_list,column_list=output_column_list)
    cs_list = create_column_scan()
    scan_df = column_check(column_list=output_column_list,column_scan_list=cs_list)
    OutputExcel(df=scan_df,file_path=excel_filename,sheet_name='柱檢核表',auto_fit_columns=[1],auto_fit_rows=[1],
                columns_list=range(2,len(scan_df.columns)+2),rows_list=range(2,len(scan_df.index)+2))
    rebar_df,concrete_df,coupler_df = summary_floor_rebar(floor_list=floor_list)
    column_df = output_col_excel(column_list=output_column_list,output_folder=output_folder,project_name=project_name)
    OutputExcel(df=rebar_df,file_path=excel_filename,sheet_name='鋼筋統計表')
    OutputExcel(df=concrete_df,file_path=excel_filename,sheet_name='混凝土統計表')
    OutputExcel(df=coupler_df,file_path=excel_filename,sheet_name='續接器統計表')
    OutputExcel(df=column_df,file_path=excel_filename,sheet_name='柱統計表')

    return excel_filename

def concat_grid_line(line_list:list,start_line:list,overlap:function):
    while True:
        temp_coor = (start_line[1],start_line[2])
        temp_line_list = [l for l in line_list if l[0] == start_line[0] and overlap(l,start_line)]
        if len(temp_line_list) == 0:
            break
        new_line_top = max(temp_line_list,key = lambda l:l[2])[2]
        new_line_bot = min(temp_line_list,key = lambda l:l[1])[1]
        if start_line[2] == new_line_top and start_line[1] == new_line_bot:
            break
        start_line[2] = new_line_top
        start_line[1] = new_line_bot
    return start_line
def concat_col_to_grid(coor_to_col_set:set,coor_to_col_line_list:list):
    def _overlap(l1,l2):
        return (l2[1] - l1[2])*(l2[2] - l1[1]) <= 0
    # Step 8. 完成col_to_line_set 格式:(col, left, right, up)
    # coor_to_col_set:((coor1,coor2),string)
    # coor_to_col_line_list:[(coor1[0], min(coor1[1], coor2[1]), max(coor1[1], coor2[1]))] y向格線
    new_coor_to_col_line_list = []
    for element in coor_to_col_set:
        coor = element[0]
        col = element[1]
        left_temp_list = [l for l in coor_to_col_line_list if l[1] <= coor[0][1] and coor[0][1] <= l[2] and l[0] <= coor[0][0]]
        right_temp_list = [l for l in coor_to_col_line_list if l[1] <= coor[0][1] and coor[0][1] <= l[2] and l[0] >= coor[0][0]]
        left_closet_line = [0,float("inf"),float("-inf")]
        right_closet_line = [0,float("inf"),float("-inf")]
        if len(left_temp_list) > 0:
            left_closet_line = list(min(left_temp_list,key = lambda l:abs(coor[0][0] - l[0])))
            left_closet_line = concat_grid_line(coor_to_col_line_list,left_closet_line,_overlap)
        if len(right_temp_list) > 0:
            right_closet_line = list(min(right_temp_list,key = lambda l:abs(coor[0][0] - l[0])))
            right_closet_line = concat_grid_line(coor_to_col_line_list,right_closet_line,_overlap)
        top = max(left_closet_line[2],right_closet_line[2])
        bot = min(left_closet_line[1],right_closet_line[1])
        new_coor_to_col_line_list.append((col,(left_closet_line[0],right_closet_line[0],bot,top)))
    return new_coor_to_col_line_list
def concat_floor_to_grid(coor_to_floor_set:set,coor_to_floor_line_list:list):
    def _overlap(l1,l2):
        return (l2[1] - l1[2])*(l2[2] - l1[1]) <= 0
    new_coor_to_floor_line_list = []
    for element in coor_to_floor_set:
        coor = element[0]
        floor = element[1]
        top_temp_list = [l for l in coor_to_floor_line_list if l[1] <= coor[0][0] and coor[0][0] <= l[2] and l[0] >= coor[0][1]]
        bot_temp_list = [l for l in coor_to_floor_line_list if l[1] <= coor[0][0] and coor[0][0] <= l[2] and l[0] <= coor[0][1]]
        top_closet_line = [0,float("inf"),float("-inf")]
        bot_closet_line = [0,float("inf"),float("-inf")]
        if len(top_temp_list) > 0:
            top_closet_line = list(min(top_temp_list,key = lambda l:abs(coor[0][1] - l[0])))
            top_closet_line = concat_grid_line(coor_to_floor_line_list,top_closet_line,_overlap)
        if len(bot_temp_list) > 0:
            bot_closet_line = list(min(bot_temp_list,key = lambda l:abs(coor[0][1] - l[0])))
            bot_closet_line = concat_grid_line(coor_to_floor_line_list,bot_closet_line,_overlap)
        right = max(top_closet_line[2],bot_closet_line[2])
        left = min(top_closet_line[1],bot_closet_line[1])
        new_coor_to_floor_line_list.append((floor,(left,right,bot_closet_line[0],top_closet_line[0])))
    return new_coor_to_floor_line_list

def draw_grid_line(new_coor_to_floor_line_list:list,new_coor_to_col_line_list:list,msp_beam:object,doc_beam:object):
    output_dwg = os.path.join(output_folder,f'{project_name}_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_Markon.dwg')
    for grid in new_coor_to_floor_line_list:
        coor_list1 = [grid[1][0], grid[1][2], 0, grid[1][1], grid[1][2], 0,grid[1][1], grid[1][3], 0,grid[1][0], grid[1][3], 0]
        # coor_list2 = [beam.coor.x, beam.coor.y, 0, rebar.end_pt.x, rebar.end_pt.y, 0]
        points1 = vtFloat(coor_list1)
        line1 = msp_beam.AddPolyline(points1)
        text1 = msp_beam.AddMText(vtPnt((grid[1][0] + grid[1][1])/2, (grid[1][2] + grid[1][3])/2),10,grid[0])
        text1.Height = 50
        line1.SetWidth(0, 5, 5)
        line1.color = 101
    doc_beam.SaveAs(output_dwg)
    doc_beam.Close(SaveChanges=True)
        # points2 = vtFloat(coor_list2)
def _ingrid(size_coor,grid_coor):
    pt_x = size_coor[0]
    pt_y = size_coor[1]
    if len(grid_coor) == 0:return False
    if (pt_x - grid_coor[0])*(pt_x - grid_coor[1])<0 and (pt_y - grid_coor[2])*(pt_y - grid_coor[3])<0:
        return True
    return False
def concat_name_to_col_floor(coor_to_size_set:set,new_coor_to_col_line_list:list,new_coor_to_floor_line_list:list,coor_to_section_list:list):

    output_column_list:list[Column]
    output_column_list = []
    for element in coor_to_size_set:
        new_column = Column()
        coor = element[0]
        size = element[1]
        new_column.set_size(size)
        col = [c for c in new_coor_to_col_line_list if _ingrid(size_coor=coor[0],grid_coor=c[1])]
        floor = [f for f in new_coor_to_floor_line_list if _ingrid(size_coor=coor[0],grid_coor=f[1])]
        if col and floor:
            new_column.set_border(col[0][1],floor[0][1])
        if len(col) > 0:
            new_column.serial = col[0][0]
            new_column.multi_column.extend(list(map(lambda c:c[0],col[0:])))
        if len(floor) > 0:
            new_column.floor = floor[0][0]
        if len(col) > 1:
            print(f'{size}:{coor} => {list(map(lambda c:c[0],col))}')
        if len(floor) > 1:
            print(f'{size}:{coor} => {list(map(lambda c:c[0],floor))}')
            new_column.multi_floor.extend(list(map(lambda c:c[0],floor[1:])))
        if new_column.serial != '':output_column_list.append(new_column)
    return output_column_list
def get_size_from_section(new_coor_to_col_line_list:list,new_coor_to_floor_line_list:list,coor_to_section_list:list,coor_to_size_set:set):
    output_column_list:list[Column]
    output_column_list = []
    size_text = ''
    for coor1,coor2 in coor_to_section_list:
        new_column = Column()
        x_size = abs(coor1[0] - coor2[0])
        y_size = abs(coor1[1] - coor2[1])
        size = f'{x_size}x{y_size}'
        col = [c for c in new_coor_to_col_line_list if _ingrid(size_coor=coor1,grid_coor=c[1])]
        floor = [f for f in new_coor_to_floor_line_list if _ingrid(size_coor=coor1,grid_coor=f[1])]
        if col and floor:
            new_column.set_border(col[0][1],floor[0][1])
            size_text = [s for s in coor_to_size_set if new_column.in_grid(coor=s[0][0])]
            if size_text:
                print(f'CAD:{size}  TEXT:{size_text[0][1]}')
                size = size_text[0][1]
            new_column.set_size(size)
        if len(col) > 0:
            new_column.serial = col[0][0]
            new_column.multi_column.extend(list(map(lambda c:c[0],col[0:])))
        if len(floor) > 0:
            new_column.floor = floor[0][0]
        if len(col) > 1:
            print(f'{size}:{coor1} => {list(map(lambda c:c[0],col))}')
        if len(floor) > 1:
            print(f'{size}:{coor1} => {list(map(lambda c:c[0],floor))}')
            new_column.multi_floor.extend(list(map(lambda c:c[0],floor[1:])))
        if len([c for c in output_column_list if c.floor == new_column.floor and c.serial == new_column.serial]) > 0:
            print(f'{new_column.floor}{new_column.serial} is exists')
            continue
        if new_column.serial != '' :output_column_list.append(new_column)
    return output_column_list
def combine_col_rebar(column_list:list[Column],coor_to_rebar_list:list,coor_to_rebar_text_list:list):
    for coor,rebar_text in coor_to_rebar_text_list:
        column = [c for c in column_list if c.in_grid(coor=coor[0])]
        if len(column) > 0:
            if (coor[0],rebar_text) in column[0].multi_rebar_text:
                print(f'{coor[0]}:{rebar_text} is exists')
                continue
            column[0].multi_rebar_text.append((coor[0],rebar_text))
            # if column[0].rebar_text == '':
            #     column[0].rebar_text = rebar_text
            #     column[0].rebar_text_coor = coor[0]
            #     column[0].multi_rebar_text.append((coor[0],rebar_text))
            # else:
            #     column[0].multi_rebar_text.append((coor[0],rebar_text))
    for rebar in coor_to_rebar_list:
        column = [c for c in column_list if c.in_grid(rebar[0])]
        if len(column) > 0:
            column[0].add_rebar_coor(rebar)
    for column in column_list:
        column.sort_rebar()
        # print(f'{column.floor}:{column.serial} x:{column.x_row} y:{column.y_row}')

def combine_col_tie(column_list:list[Column],coor_to_tie_text_list:list,coor_to_tie_list:list):
    for tie in coor_to_tie_list:
        column = [c for c in column_list if c.in_grid(coor=tie[0]) and c.in_grid(coor=tie[1])]
        if len(column) > 0:
            column[0].add_tie(tie)
    for coor,tie_text in coor_to_tie_text_list:
        column = [c for c in column_list if c.in_grid(coor=coor[1]) and c.in_grid(coor=coor[1])]
        if len(column) > 0:
            column[0].add_tie_text(coor=coor,text=tie_text)
    for column in column_list:
        column.sort_tie()
        # print(f'{column.floor}:{column.serial} x:{column.x_tie} y:{column.y_tie} tie:{column.tie_dict}')


def output_col_excel(column_list:list[Column],output_folder:str,project_name:str):
    header_info_1 = [('樓層', ''), ('柱編號', ''), ('X向 柱寬', 'cm'), ('Y向 柱寬', 'cm')]
    header_rebar = [('柱主筋', '主筋'),('柱主筋', 'X向支數'),('柱主筋', 'Y向支數'),('柱箍筋', '圍束區'),('柱箍筋', '非圍束區'),('柱箍筋', 'X向繫筋'),('柱箍筋', 'Y向繫筋')]
    header_second_rebar = [('次柱主筋', '主筋'),('次柱主筋', 'X向支數'),('次柱主筋', 'Y向支數')]
    sorted(column_list,key = lambda c:c.serial)
    header = pd.MultiIndex.from_tuples(header_info_1 + header_rebar)
    column_df = pd.DataFrame(np.empty([len(column_list),len(header)],dtype='<U16'),columns=header)
    row = 0
    for c in column_list:
        if c.serial == '': continue
        column_df.at[row,('樓層', '')] = c.floor
        column_df.at[row,('柱編號', '')] = c.serial
        column_df.at[row,('X向 柱寬', 'cm')] = c.x_size
        column_df.at[row,('Y向 柱寬', 'cm')] = c.y_size
        if len(c.total_rebar) > 0: 
            column_df.at[row,('柱主筋', '主筋')] = c.total_rebar[0][0].text
            column_df.at[row,('柱主筋', 'X向支數')] = c.x_dict[c.total_rebar[0][0].size]
            column_df.at[row,('柱主筋', 'Y向支數')] = c.y_dict[c.total_rebar[0][0].size]
        if len(c.total_rebar) == 2:
            column_df.at[row,('次柱主筋', '主筋')] = c.total_rebar[1][0].text
            column_df.at[row,('次柱主筋', 'X向支數')] = c.x_dict[c.total_rebar[1][0].size]
            column_df.at[row,('次柱主筋', 'Y向支數')] = c.y_dict[c.total_rebar[1][0].size]
        if c.tie_dict:
            column_df.at[row,('柱箍筋', '圍束區')] = c.tie_dict['端部'][1]
            column_df.at[row,('柱箍筋', '非圍束區')] = c.tie_dict['中央'][1]
            column_df.at[row,('柱箍筋', 'X向繫筋')] = c.x_tie
            column_df.at[row,('柱箍筋', 'Y向繫筋')] = c.y_tie
        row += 1
    
    # output_column_list = sorted(output_column_list,key=lambda c:c.serial)
    # column_df.sort_values(by=[('柱編號', '')],ascending=True,inplace=True)
    return column_df
def floor_parameter(column_list:list[Column],floor_parameter_xlsx:str):
    floor_list:list[Floor]
    floor_list = []
    parameter_df = read_parameter_df(floor_parameter_xlsx)
    parameter_df.set_index(['樓層'],inplace=True)
    for c in column_list:
        for floor in c.multi_floor:
           for column_name in c.multi_column:
                new_c = copy.deepcopy(c)
                new_c.floor = floor
                new_c.serial = column_name
                new_c.multi_floor = []
                column_list.append(new_c) 
    for floor_name in parameter_df.index:
        temp_floor = Floor(floor_name)
        floor_list.append(temp_floor)
        temp_floor.set_prop(parameter_df.loc[floor_name])
        temp_floor.add_column([c for c in column_list if c.floor == temp_floor.floor_name])
    
    return floor_list
def sort_floor_column(floor_list:list[Floor],column_list:list[Column]):
    def match_column(col:Column,col_list:list[Column],pos:str):
        temp_list = [c for c in col_list if c.serial == col.serial]
        if temp_list and pos == 'up':
            col.up_column = temp_list[0]
        if temp_list and pos == 'bot':
            col.bot_column = temp_list[0]
    floor_seq = list(map(lambda f:f.floor_name,floor_list))
    list(map(lambda c:c.set_seq(floor_seq),column_list))
    for i in range(0,len(floor_list) - 1):
        temp_list = floor_list[i].column_list
        bot_list = floor_list[i + 1].column_list
        list(map(lambda c:match_column(c,bot_list,'bot'),temp_list))
    for i in range(1,len(floor_list)):
        temp_list = floor_list[i].column_list
        up_list = floor_list[i - 1].column_list
        list(map(lambda c:match_column(c,up_list,'up'),temp_list))
    column_list.sort(key=lambda c:(c.serial,-1*c.seq))
def summary_floor_rebar(floor_list:list[Floor]):
    df = pd.DataFrame(columns=['#3','#4','#5','#6','#7','#8','#10','#11'],index=[])
    concrete_df = pd.DataFrame(columns=[],index=[])
    coupler_df = pd.DataFrame(columns=[],index=[])
    for floor in floor_list:
        list(map(lambda c:c.calculate_rebar() ,floor.column_list))
        floor.summary_rebar()
        new_row = pd.DataFrame(floor.rebar_count,index=[floor.floor_name])
        new_row_concrete = pd.DataFrame(floor.concrete_count,index=[floor.floor_name])
        new_row_coupler = pd.DataFrame(floor.coupler,index=[floor.floor_name])
        df = pd.concat([df, new_row], verify_integrity=True)
        concrete_df = pd.concat([concrete_df,new_row_concrete],verify_integrity=True)
        coupler_df = pd.concat([coupler_df,new_row_coupler],verify_integrity=True)
    df.fillna(value=0,inplace=True)
    df.loc['Sum'] = df.sum()
    concrete_df.loc['Sum'] = concrete_df.sum()
    return df,concrete_df,coupler_df    
def read_parameter_df(read_file):
    return pd.read_excel(
        read_file, sheet_name='參數表',header=[0])

def OutputExcel(df:pd.DataFrame,file_path,sheet_name,auto_fit_columns=[],auto_fit_rows=[],columns_list=[],rows_list=[]):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl') 
        writer.book = book
        # sheet = book[sheet_name]
        # sheet.column_dimensions['A'] =ColumnDimension(sheet,'L',bestFit=True)
    else:
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter') 
    df.to_excel(writer,sheet_name=sheet_name)
    writer.save()

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    writer.book = book
    if os.path.exists(file_path) and len(auto_fit_columns) >0:
        AutoFit_Columns(book[sheet_name],auto_fit_columns,auto_fit_rows)
    if os.path.exists(file_path) and len(columns_list) >0:
        Decorate_Worksheet(book[sheet_name],columns_list,rows_list)
    writer.save()
    return file_path

def Decorate_Worksheet(sheet:Worksheet,columns_list:list,rows_list:list):
    for i in columns_list:
        for j in rows_list:
            sheet.cell(j,i).alignment = Alignment(vertical='center',wrap_text=True,horizontal='center')
            sheet.cell(j,i).font = Font(name='Calibri')
            if sheet.cell(j,i).value == 'NG.':sheet.cell(j,i).fill = PatternFill("solid",start_color='00FF0000')

def AutoFit_Columns(sheet:Worksheet,auto_fit_columns:list,auto_fit_rows:list):
    for i in auto_fit_columns:
        sheet.column_dimensions[get_column_letter(i)].width = 80
    for i in auto_fit_rows:
        sheet.row_dimensions[i].height = 20
    for i in auto_fit_rows:
        for j in auto_fit_columns:
            sheet.cell(i,j).alignment = Alignment(wrap_text=True,vertical='center',horizontal='center')

def count_column_multiprocessing(column_filenames:list[str]):
    coor_to_col_set = ()
    coor_to_size_set = ()
    coor_to_floor_set = ()
    coor_to_col_line_list = []
    coor_to_floor_line_list = []
    coor_to_rebar_text_list = []
    coor_to_rebar_list = []
    coor_to_tie_text_list = []
    coor_to_tie_list = []
    coor_to_section_list = []
    with Pool(processes=10) as p:
        start = time.time() # 開始測量執行時間
        jobs = []
        for filename in column_filenames:
            jobs.append(p.apply_async(read_column_cad, (filename,)))
        for job in jobs:

def count_column_main(column_filename,layer_config,temp_file='temp_1221_1F.pkl',output_folder='',project_name='',template_name='',floor_parameter_xlsx = ''):
    progress_file = './result/tmp'
    start = time.time()
    msp_column,doc_column = read_column_cad(beam_filename=column_filename,progress_file=progress_file)
    sort_col_cad(msp_beam=msp_column,layer_config=layer_config,temp_file=temp_file)
    output_excel = cal_column_rebar(data=save_temp_file.read_temp(temp_file),output_folder=output_folder,
                                    project_name=project_name,progress_file=progress_file,floor_parameter_xlsx = floor_parameter_xlsx)
    # output_dwg = draw_rebar_line(class_beam_list=class_beam_list,msp_beam=msp_column,doc_beam=doc_column,output_folder=output_folder,project_name=project_name)
    print(f'Total Time:{time.time() - start}')
    return os.path.basename(output_excel)
if __name__ == '__main__':
    col_filename = r'D:\Desktop\BeamQC\TEST\2023-0203\圓方烏日-XS-COL.dwg'#sys.argv[1] # XS-COL的路徑
    floor_parameter_xlsx = r'D:\Desktop\BeamQC\TEST\2023-0203\'
    output_folder ='D:/Desktop/BeamQC/TEST/OUTPUT/'
    project_name = 'test_column'
    # layer_config = {
    #     'text_layer':['TABLE','SIZE'],
    #     'line_layer':['TABLE'],
    #     'rebar_text_layer':['NBAR'], # 箭頭和鋼筋文字的塗層
    #     'rebar_layer':['RBAR'], # 鋼筋和箍筋的線的塗層
    #     'tie_text_layer':['NBAR'], # 箍筋文字圖層
    #     'tie_layer':['RBAR'], # 箍筋文字圖層
    #     'block_layer':['DwFm'], # 框框的圖層
    #     'column_rc_layer':['OLINE'] #斷面圖層
    # }
    #DrawRC
    entity_type ={
        'rebar_layer':['AcDbPolyline'],
        'rebar_data_layer':['AcDbMText'],
        'rebar_data_leader_layer':['AcDbLeader'],
        'tie_text_layer':['AcDbText']
    }
    #RCAD
    # layer_config = {
    #     'text_layer':['文字-柱線名稱','文字-樓群名稱','文字-斷面尺寸'],
    #     'line_layer':['GirdInner'],
    #     'rebar_text_layer':['文字-主筋根數'], # 箭頭和鋼筋文字的塗層
    #     'rebar_layer':['主筋斷面'], # 鋼筋和箍筋的線的塗層
    #     'tie_text_layer':['文字-剪力筋 中央區','文字-剪力筋-BC','文字-剪力筋-圍束區'], # 箍筋文字圖層
    #     'tie_layer':['箍筋線'], # 箍筋文字圖層
    #     'block_layer':['DEFPOINTS'], # 框框的圖層
    #     'column_rc_layer':['柱斷面線'] #斷面圖層
    # }
    #Elements
    layer_config = {
        'text_layer':['S-TEXT'],
        'line_layer':['S-STUD'],
        'rebar_text_layer':['S-TEXT'], # 箭頭和鋼筋文字的塗層
        'rebar_layer':['S-REINFD'], # 鋼筋和箍筋的線的塗層
        'tie_text_layer':['S-TEXT'], # 箍筋文字圖層
        'tie_layer':['S-REINF'], # 箍筋文字圖層
        'block_layer':['Page'], # 框框的圖層
        'column_rc_layer':['S-RC'] #斷面圖層
    }
    msp_column = None
    doc_column = None
    # msp_column,doc_column = read_column_cad(col_filename,layer_config)
    # sort_col_cad(msp_column=msp_column,layer_config=layer_config,temp_file='temp_col_Elements_ALL_0213.pkl')
    cal_column_rebar(data=save_temp_file.read_temp(r'temp_col_Elements_ALL_0213.pkl'),output_folder=output_folder,project_name=project_name,msp_column= msp_column,doc_column= doc_column)
    # floor_list = floor_parameter(column_list)
    # coor_to_col_set = set()
    # coor_to_col_set.add((((0,0),(10,10)),"C1"))
    # coor_to_col_line_list = [(-5,0,10),(-5,0,10),(-5,5,15),(-5,15,20),(-5,18,25),(-5,-5,30),(-5,-5,30),(10,0,10)]
    # print(temp(coor_to_col_set,coor_to_col_line_list))
    # temp = [[1,2],[3,4]]
    # print(f'{list(map(lambda t:t[0],temp))}')
    # print(pd.DataFrame({'#1':0,"#2":10},index=[]))
    # print(re.search(r'(\D+)\d+(\D+)','C15CC').group(2))
    # temp = save_temp_file.read_temp(r'temp_col_Elements_0213.pkl')
    # print(temp)