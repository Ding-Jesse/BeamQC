from __future__ import annotations
from multiprocessing.spawn import prepare
import time
import multiprocessing
# from AutocadApi import read_plan,read_beam,write_beam,write_plan,error
from plan_to_beam import run_plan,run_beam,output_error_list,write_beam,write_plan,write_result_log
from werkzeug.utils import secure_filename
import os
import plan_to_col
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.worksheet.worksheet import Worksheet

def main_functionV3(beam_filenames,plan_filenames,beam_new_filename,plan_new_filename,project_name,output_directory,layer_config,progress_file,sizing,mline_scaling):
    start = time.time()

    plan_file = './result/plan.txt' # plan.txt的路徑
    beam_file = './result/beam.txt' # beam.txt的路徑
    excel_file = './result/result_log.xlsx' # result_log.xlsx的路徑
    big_file = os.path.join(output_directory,f'{project_name}-大梁.txt')
    sml_file = os.path.join(output_directory,f'{project_name}-小梁.txt')
    fbeam_file = os.path.join(output_directory,f'{project_name}-地梁.txt')
    task_name = project_name
    date = time.strftime("%Y-%m-%d", time.localtime())

    # layer_config = {
    #     'text_layer':text_layer,
    #     'block_layer':block_layer,
    #     'floor_layer':floor_layer,
    #     'big_beam_layer':big_beam_layer,
    #     'big_beam_text_layer':big_beam_text_layer,
    #     'sml_beam_layer':sml_beam_layer,
    #     'size_layer':size_layer,
    #     'sml_beam_text_layer':sml_beam_text_layer
    # }

    multiprocessing.freeze_support()    
    pool = multiprocessing.Pool()

    res_plan =[]
    res_beam = []
    set_plan = set()
    dic_plan = {}
    set_beam = set()
    dic_beam = {}

    for plan_filename in plan_filenames:
        res_plan.append(pool.apply_async(run_plan,(plan_filename, plan_new_filename, big_file, sml_file,layer_config , plan_file, progress_file, sizing, mline_scaling, date,fbeam_file)))
    for beam_filename in beam_filenames:
        res_beam.append(pool.apply_async(run_beam,(beam_filename, layer_config['text_layer'], beam_file, progress_file, sizing)))
        
    plan_drawing = 0
    if len(plan_filenames) == 1:
        plan_drawing = 1
    beam_drawing = 0
    if len(beam_filenames) == 1:
        beam_drawing = 1

    for plan in res_plan:
        plan = plan.get()
        if plan:
            set_plan = set_plan | plan[0]
            if plan_drawing:
                dic_plan = plan[1]
        else:
            end = time.time()
            write_result_log(excel_file = excel_file, 
                    task_name = task_name, 
                    plan_result=[], 
                    beam_result=[], 
                    runtime = f'{round(end - start, 2)}(s)', 
                    date= time.strftime("%Y-%m-%d %H:%M", time.localtime()), 
                    other='plan failed')
            return

    for beam in res_beam:
        beam = beam.get()
        if beam:
            set_beam = set_beam | beam[0]
            if beam_drawing:
                dic_beam = beam[1]
        else:
            end = time.time()
            write_result_log(excel_file = excel_file, 
                             task_name = task_name, 
                             plan_result=[], 
                             beam_result=[], 
                             runtime = f'{round(end - start, 2)}(s)', 
                             date= time.strftime("%Y-%m-%d %H:%M", time.localtime()), 
                             other='beam failed')
            return

    plan_error_list,f_fbeam,f_big,f_sml = write_plan(plan_filename, 
                                                     plan_new_filename, 
                                                     set_plan, set_beam, 
                                                     dic_plan, big_file, 
                                                     sml_file, date, 
                                                     plan_drawing, progress_file, 
                                                     sizing, mline_scaling,
                                                     fbeam_file=fbeam_file)
    plan_result = output_error_list(error_list=plan_error_list,
                                    f_fbeam=f_fbeam,
                                    f_big=f_big,
                                    f_sml=f_sml,
                                    title_text='XS-BEAM',
                                    set_item=set_plan,
                                    progress_file=progress_file)
    beam_error_list,f_fbeam,f_big,f_sml = write_beam(beam_filename, 
                                                     beam_new_filename, 
                                                     set_plan, 
                                                     set_beam, 
                                                     dic_beam, 
                                                     big_file, 
                                                     sml_file, 
                                                     date, 
                                                     beam_drawing, 
                                                     progress_file, 
                                                     sizing,
                                                     fbeam_file=fbeam_file)
    beam_result = output_error_list(error_list=beam_error_list,
                                    f_sml=f_sml,
                                    f_big=f_big,
                                    f_fbeam=f_fbeam,
                                    title_text='XS-PLAN',
                                    set_item=set_beam,
                                    progress_file=progress_file)
    end = time.time()
    write_result_log(excel_file = excel_file,
                     task_name= task_name,
                     plan_result = plan_result,
                     beam_result= beam_result,
                     runtime= f'{round(end - start, 2)}(s)',
                     date = time.strftime("%Y-%m-%d %H:%M", time.localtime()),
                     other =  'none')
    return [os.path.basename(big_file),os.path.basename(sml_file),os.path.basename(fbeam_file)]
def main_col_function(col_filenames,plan_filenames,col_new_filename,plan_new_filename,result_file,text_layer,line_layer,block_layer,floor_layer,col_layer,task_name,progress_file):
    start = time.time()

    plan_file = './result/col_plan.txt' # plan.txt的路徑
    col_file = './result/col.txt' # col.txt的路徑
    excel_file = './result/result_log_col.xlsx' # result_log.xlsx的路徑

    date = time.strftime("%Y-%m-%d", time.localtime())

    multiprocessing.freeze_support()
    pool = multiprocessing.Pool()
    res_plan =[]
    res_col = []
    set_plan = set()
    dic_plan = {}
    set_col = set()
    dic_col = {}
    for plan_filename in plan_filenames:
        res_plan.append(pool.apply_async(plan_to_col.read_plan, (plan_filename, floor_layer, col_layer, block_layer, plan_file, progress_file)))
    for col_filename in col_filenames:
        res_col.append(pool.apply_async(plan_to_col.read_col, (col_filename, text_layer, line_layer, col_file, progress_file)))

    plan_drawing = 0
    if len(plan_filenames) == 1:
        plan_drawing = 1
    col_drawing = 0
    if len(col_filenames) == 1:
        col_drawing = 1

    for plan in res_plan:
        plan = plan.get()
        if plan:
            set_plan = set_plan | plan[0]
            if plan_drawing:
                dic_plan = plan[1]
        else:
            end = time.time()
            plan_to_col.write_result_log(excel_file,task_name,'','', f'{round(end - start, 2)}(s)', time.strftime("%Y-%m-%d %H:%M", time.localtime()), 'failed')
            return 

    for col in res_col:
        col = col.get()
        if col:
            set_col = set_col | col[0]
            if col_drawing:
                dic_col = col[1]
        else:
            end = time.time()
            plan_to_col.write_result_log(excel_file,task_name,'','', f'{round(end - start, 2)}(s)', time.strftime("%Y-%m-%d %H:%M", time.localtime()), 'failed')
            return 
                
    plan_result = plan_to_col.write_plan(plan_filename, plan_new_filename, set_plan, set_col, dic_plan, result_file, date, plan_drawing, progress_file)
    col_result = plan_to_col.write_col(col_filename, col_new_filename, set_plan, set_col, dic_col, result_file, date,col_drawing, progress_file)

    end = time.time()
    plan_to_col.write_result_log(excel_file,task_name,plan_result,col_result, f'{round(end - start, 2)}(s)', time.strftime("%Y-%m-%d %H:%M", time.localtime()), 'none')
    return 

def storefile(file,file_directory,file_new_directory,project_name):
    filename_beam = secure_filename(file.filename)
    check_beam_name = os.path.splitext(filename_beam)
    if check_beam_name[1] == '':
        filename_beam = f'temp.{check_beam_name[0]}'
    print(f'file:{file.filename} file:{filename_beam}')
    save_file = os.path.join(file_directory, f'{project_name}-{filename_beam}')
    file_new_name = os.path.join(file_new_directory, f'{project_name}_MARKON-{filename_beam}')
    file.save(save_file)
    file_ok = True
    return file_ok , file_new_name,save_file

def OutputExcel(df_list:list[pd.DataFrame],file_path,sheet_name,auto_fit_columns=[],auto_fit_rows=[],columns_list=[],rows_list=[],df_spacing = 0 ):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl') 
        writer.book = book
        # sheet = book[sheet_name]
        # sheet.column_dimensions['A'] =ColumnDimension(sheet,'L',bestFit=True)
    else:
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    row = 0
    for df in df_list: 
        df.to_excel(writer,sheet_name=sheet_name,startrow=row)
        row += len(df.index) + df_spacing
    writer.save()

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    writer.book = book
    if os.path.exists(file_path) and len(auto_fit_columns) >0:
        AutoFit_Columns(book[sheet_name],auto_fit_columns,auto_fit_rows)
    if os.path.exists(file_path) and len(columns_list) >0:
        Decorate_Worksheet(book[sheet_name],columns_list,rows_list)
    writer.save()
    return file_path

def Decorate_Worksheet(sheet:Worksheet,columns_list:list,rows_list:list):
    for i in columns_list:
        for j in rows_list:
            sheet.cell(j,i).alignment = Alignment(vertical='center',wrap_text=True,horizontal='center')
            sheet.cell(j,i).font = Font(name='Calibri')
            if sheet.cell(j,i).value == 'NG.':sheet.cell(j,i).fill = PatternFill("solid",start_color='00FF0000')

def AutoFit_Columns(sheet:Worksheet,auto_fit_columns:list,auto_fit_rows:list):
    for i in auto_fit_columns:
        sheet.column_dimensions[get_column_letter(i)].width = 80
    for i in auto_fit_rows:
        sheet.row_dimensions[i].height = 20
    for i in auto_fit_rows:
        for j in auto_fit_columns:
            sheet.cell(i,j).alignment = Alignment(wrap_text=True,vertical='center',horizontal='center')

def Add_Row_Title(file_path:str,sheet_name:str,i:int,j:int,title_text:str,font_size = 12):
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    writer.book = book
    sheet = book[sheet_name]
    sheet.cell(i,j).value = title_text
    sheet.cell(i,j).alignment = Alignment(vertical='center',wrap_text=True,horizontal='center')
    sheet.cell(i,j).font = Font(name='Calibri',size= font_size)
    writer.save()
def Output_Config(project_name:str,layer_config:dict,file_new_directory:str):
    with open(os.path.join(file_new_directory, f'{project_name}_layer_config.txt'),'w') as f:
        f.write(str(layer_config))
    pass
def GetAllFiles(mypath:str):
    import glob
    from os.path import isfile, join
    return glob.glob(join(mypath,"*.dwg"))
if __name__ == '__main__':
    floor_text = 'R1F-13F'
    stash_pattern = r'(\w+)[-](\w+)'
    import re
    from numpy import arange
    from plan_to_beam import turn_floor_to_float,turn_floor_to_string
    def _get_floor_list(floor1:float,floor2:float):
        if floor1 >= floor2:
            l = list(range(int(floor1),int(floor2),-1))
            l.append(floor2)
            return l
        else:
            l = list(range(int(floor1),int(floor2),1))
            l.append(floor2)
            return l
    floor_tuple = re.findall(stash_pattern ,floor_text)
    for floors in floor_tuple:
        first_floor = turn_floor_to_float(floor=floors[0])
        second_floor = turn_floor_to_float(floor=floors[-1])
        if first_floor and second_floor and max(first_floor,second_floor) < 100:
            for floor_float in _get_floor_list(second_floor,first_floor):
                print(turn_floor_to_string(floor_float))
            print('1')
    # import glob
    # from os.path import isfile, join
    # mypath = r'D:\Desktop\BeamQC\TEST\2023-0320\東仁'
    # print(type(glob.glob(join(mypath,"*.dwg"))))
    # beam_filenames = [r'D:\Desktop\BeamQC\TEST\2023-0310\XS-BEAM(南基地).dwg']
    # plan_filenames = [r'D:\Desktop\BeamQC\TEST\2023-0310\岡山(南基地)-XS-PLAN-TEST.dwg']#sys.argv[2] # XS-PLAN的路徑
    # beam_new_filename = r"D:\Desktop\BeamQC\TEST\XS-BEAM_new.dwg"#sys.argv[3] # XS-BEAM_new的路徑
    # plan_new_filename = r"D:\Desktop\BeamQC\TEST\XS-PLAN_new.dwg"#sys.argv[4] # XS-PLAN_new的路徑
    # big_file = r"D:\Desktop\BeamQC\TEST\big-4.txt"#sys.argv[5] # 大梁結果
    # sml_file = r"D:\Desktop\BeamQC\TEST\sml-4.txt"#sys.argv[6] # 小梁結果
    # fbeam_file = r"D:\Desktop\BeamQC\TEST\fb-4.txt"#sys.argv[6] # 地梁結果
    # # 在beam裡面自訂圖層
    # text_layer = 'S-RC'#sys.argv[7]

    # # 在plan裡面自訂圖層
    # block_layer = 'DwFm'#sys.argv[8] # 框框的圖層
    # floor_layer = 'S-TITLE'#sys.argv[9] # 樓層字串的圖層
    # size_layer = 'S-TEXT'#sys.argv[12] # 梁尺寸字串圖層
    # big_beam_layer = 'S-RCBMG'#大樑複線圖層
    # big_beam_text_layer = 'S-TEXTG'#大樑文字圖層
    # sml_beam_layer = 'S-RCBMB'#小梁複線圖層
    # sml_beam_text_layer = 'S-TEXTB'#小梁文字圖層
    # task_name = 'temp'#sys.argv[13]

    # progress_file = './result/tmp'#sys.argv[14]

    # sizing = 1 # 要不要對尺寸
    # mline_scaling = 1 # 要不要對複線寬度

    # plan_file = './result/plan.txt' # plan.txt的路徑
    # beam_file = './result/beam.txt' # beam.txt的路徑
    # excel_file = './result/result_log.xlsx' # result_log.xlsx的路徑
    
    # date = time.strftime("%Y-%m-%d", time.localtime())
    # layer_config = {
    #     # 'line_layer':line_layer,
    #     'text_layer':text_layer,
    #     'block_layer':block_layer,
    #     'floor_layer':floor_layer,
    #     'big_beam_layer':big_beam_layer,
    #     'big_beam_text_layer':big_beam_text_layer,
    #     'sml_beam_layer':sml_beam_layer,
    #     'size_layer':size_layer,
    #     'sml_beam_text_layer':sml_beam_text_layer
    #     # 'col_layer':col_layer
    # }
    # # from collections import Counter
    # main_functionV3(beam_filenames=beam_filenames,
    #                 beam_new_filename=beam_new_filename,
    #                 plan_filenames=plan_filenames,
    #                 plan_new_filename=plan_new_filename,
    #                 project_name=task_name,
    #                 output_directory=r"D:\Desktop\BeamQC\TEST\2023-0310",
    #                 layer_config=layer_config,
    #                 progress_file=progress_file,
    #                 sizing=sizing,
    #                 mline_scaling=mline_scaling)
    # counter.update({'A':1,'B':1})
    # counter.update({'A':1})
    # counter.update({'A':1})
    # counter.keys()
    # # data={
    # #     'A':1,
    # #     'B':2
    # # }
    # df = pd.DataFrame.from_dict(counter, orient='index',columns=['item'])
    # # df = pd.DataFrame(data=data,columns=['item'],index=data.keys())
    # print(df)
    # from functools import wraps
    # def logger(function):
    #     # @wraps(function)
    #     def wrapper(*args, **kwargs):
    #         """wrapper documentation"""
    #         print(f"----- {function.__name__}: start -----")
    #         output = function(*args, **kwargs)
    #         print(f"----- {function.__name__}: end -----")
    #         return output
    #     return wrapper
    # # @logger
    # def f1(v):
    #     print('f1')
    #     # @logger
    #     def f2(v):
    #         print(v)
    #     def f3(v):
    #         return f2(v=v)
    #     return f3(v)
    # print(f1(10))